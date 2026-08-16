from __future__ import annotations

import gzip
import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from visaradar.ingest import LCARow, aggregate

RAW_FILES = {
    "2024": "LCA_Disclosure_Data_FY2024_Q4.xlsx",
    "2025": "LCA_Disclosure_Data_FY2025_Q4.xlsx",
    "2026": "LCA_Disclosure_Data_FY2026_Q3.xlsx",
}

VISA_CLASSES = {"H-1B", "H-1B1", "E-3"}


def read_rows(path: Path, fiscal_year: str) -> list[LCARow]:
    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = {name: i for i, name in enumerate(header)}

    required = ["EMPLOYER_NAME", "CASE_STATUS", "JOB_TITLE", "VISA_CLASS"]
    for col in required:
        if col not in idx:
            raise ValueError(f"{path.name}: missing expected column {col}")

    rows: list[LCARow] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        visa_class = values[idx["VISA_CLASS"]]
        if visa_class not in VISA_CLASSES:
            continue
        employer_name = values[idx["EMPLOYER_NAME"]]
        case_status = values[idx["CASE_STATUS"]]
        job_title = values[idx["JOB_TITLE"]]
        if not employer_name or not case_status:
            continue
        rows.append(
            LCARow(
                employer_name=str(employer_name),
                case_status=str(case_status),
                fiscal_year=fiscal_year,
                job_title=str(job_title) if job_title else "",
            )
        )
    wb.close()
    return rows


def main() -> None:
    raw_dir = Path(__file__).resolve().parent.parent / ".rawdata"
    out_path = Path(__file__).resolve().parent.parent / "visaradar" / "lca_snapshot.jsonl.gz"

    all_rows: list[LCARow] = []
    for fy, filename in RAW_FILES.items():
        path = raw_dir / filename
        if not path.exists():
            raise FileNotFoundError(f"expected raw file not found: {path}")
        print(f"reading {filename} (FY{fy})...", file=sys.stderr)
        rows = read_rows(path, fy)
        print(f"  {len(rows)} LCA rows (H-1B/H-1B1/E-3)", file=sys.stderr)
        all_rows.extend(rows)

    print(f"aggregating {len(all_rows)} total rows...", file=sys.stderr)
    aggregated = aggregate(all_rows)
    print(f"  {len(aggregated)} unique employers", file=sys.stderr)

    tmp_path = out_path.with_suffix(out_path.suffix + ".tmp")
    with gzip.open(tmp_path, "wt", encoding="utf-8") as f:
        for record in aggregated.values():
            f.write(json.dumps(record) + "\n")
    tmp_path.replace(out_path)

    size_mb = out_path.stat().st_size / (1024 * 1024)
    print(f"wrote {out_path} ({size_mb:.2f} MB)", file=sys.stderr)


if __name__ == "__main__":
    main()
